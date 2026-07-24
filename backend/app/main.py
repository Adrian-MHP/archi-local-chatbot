from __future__ import annotations

from datetime import datetime, timezone
import io
import json
import os
from queue import Empty, Queue
import re
from threading import Thread
from typing import Any, Dict
from uuid import uuid4

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import httpx
from pydantic import ValidationError

from .azure_agent import ChatService
from .config import get_settings
from .schemas import (
    ActionResponse,
    ApplyPlanRequest,
    AutomationPlanResponse,
    ChatRequest,
    ChatResponse,
    ChatStreamRequest,
    ChatTraceResponse,
    ConversationArchive,
    ConversationExportRequest,
    ConversationExportResponse,
    ConversationImportRequest,
    ConversationImportResponse,
)

load_dotenv()

settings = get_settings()
chat_service = ChatService(settings)

app = FastAPI(title="Archi Local Chatbot API", version="0.1.0")

allowed_origins = os.getenv("ALLOWED_ORIGINS", "*")
origins = [o.strip() for o in allowed_origins.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _history_from_payload(payload: ChatRequest) -> list[dict[str, str]]:
    return [{"role": item.role, "content": item.content} for item in payload.history]


def _sse(event: str, data: Dict[str, Any]) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


def _chunk_text(text: str, chunk_size: int) -> list[str]:
    if not text:
        return []
    if chunk_size <= 0:
        return [text]

    chunks: list[str] = []
    i = 0
    n = len(text)
    while i < n:
        end = min(i + chunk_size, n)
        if end < n:
            last_space = text.rfind(" ", i, end)
            if last_space > i + 20:
                end = last_space
        chunk = text[i:end].strip()
        if chunk:
            chunks.append(chunk)
        i = max(end, i + 1)
    return chunks or [text]


def _normalize_conversation(conversation: ConversationArchive) -> tuple[ConversationArchive, list[str]]:
    warnings: list[str] = []
    now = _now_iso()

    if not conversation.created_at:
        conversation.created_at = now
        warnings.append("created_at was missing and has been set automatically.")
    if not conversation.updated_at:
        conversation.updated_at = now
        warnings.append("updated_at was missing and has been set automatically.")
    if conversation.updated_at and conversation.created_at and conversation.updated_at < conversation.created_at:
        conversation.updated_at = conversation.created_at
        warnings.append("updated_at was earlier than created_at and has been corrected.")
    return conversation, warnings


def _normalize_extracted_text(text: str) -> str:
    cleaned = (text or "").replace("\x00", " ")
    cleaned = cleaned.replace("\r\n", "\n").replace("\r", "\n")
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _extract_text_from_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"PDF support is unavailable: {exc}") from exc

    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"Unable to read PDF: {exc}") from exc

    text_parts: list[str] = []
    max_pages = 50
    for page_index, page in enumerate(reader.pages):
        if page_index >= max_pages:
            break
        page_text = page.extract_text() or ""
        if page_text.strip():
            text_parts.append(page_text)
    combined = "\n\n".join(text_parts)
    if not combined.strip():
        raise HTTPException(
            status_code=422,
            detail="No readable text found in PDF. If this is a scanned PDF, OCR is required first.",
        )
    return combined


def _extract_text_from_xlsx(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Excel support is unavailable: {exc}") from exc

    try:
        workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"Unable to read Excel file: {exc}") from exc

    lines: list[str] = []
    max_sheets = 8
    max_rows_per_sheet = 500
    max_cells_per_row = 30

    try:
        for sheet_index, sheet in enumerate(workbook.worksheets):
            if sheet_index >= max_sheets:
                break
            lines.append(f"Sheet: {sheet.title}")
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index > max_rows_per_sheet:
                    break
                values: list[str] = []
                for value in row[:max_cells_per_row]:
                    if value is None:
                        continue
                    string_value = str(value).strip()
                    if string_value:
                        values.append(string_value)
                if values:
                    lines.append(" | ".join(values))
    finally:
        workbook.close()

    combined = "\n".join(lines)
    if not combined.strip():
        raise HTTPException(status_code=422, detail="No readable rows found in Excel file.")
    return combined


async def _extract_text_from_upload(file: UploadFile) -> tuple[str, str]:
    filename = (file.filename or "upload").strip() or "upload"
    lower_name = filename.lower()
    suffix = f".{lower_name.rsplit('.', 1)[-1]}" if "." in lower_name else ""

    content = await file.read()
    size = len(content)
    if size <= 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    if size > settings.max_upload_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Uploaded file exceeds size limit of {settings.max_upload_bytes} bytes.",
        )

    if suffix in {".txt", ".md", ".csv", ".json"}:
        try:
            raw_text = content.decode("utf-8")
        except UnicodeDecodeError:
            raw_text = content.decode("latin-1", errors="ignore")
    elif suffix == ".pdf":
        raw_text = _extract_text_from_pdf(content)
    elif suffix in {".xlsx", ".xlsm"}:
        raw_text = _extract_text_from_xlsx(content)
    elif suffix == ".xls":
        raise HTTPException(
            status_code=415,
            detail="Legacy .xls is not supported. Please upload .xlsx or export to PDF/text.",
        )
    else:
        raise HTTPException(
            status_code=415,
            detail="Unsupported file type. Use PDF, XLSX, XLSM, TXT, MD, CSV, or JSON.",
        )

    text = _normalize_extracted_text(raw_text)
    if not text:
        raise HTTPException(status_code=422, detail="No processable text could be extracted from uploaded file.")
    if len(text) > settings.max_upload_text_chars:
        text = text[: settings.max_upload_text_chars]
    return text, filename


@app.get("/api/health")
def health() -> Dict[str, Any]:
    mcp_status = "ok"
    mcp_error = None
    try:
        tool_count = len(chat_service.mcp.list_tools())
    except Exception as exc:  # noqa: BLE001
        mcp_status = "error"
        mcp_error = str(exc)
        tool_count = 0

    return {
        "status": "ok",
        "azure_model": settings.azure_openai_model,
        "mcp_server_url": settings.mcp_server_url,
        "mcp_status": mcp_status,
        "mcp_error": mcp_error,
        "mcp_tool_count": tool_count,
    }


@app.get("/api/tools")
def tools() -> Dict[str, Any]:
    try:
        mcp_tools = chat_service.mcp.list_tools()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Failed to list MCP tools: {exc}") from exc

    return {
        "count": len(mcp_tools),
        "tools": [
            {
                "name": t.name,
                "description": t.description,
                "input_schema": t.input_schema,
            }
            for t in mcp_tools
        ],
    }


@app.get("/api/debug/mcp")
def debug_mcp() -> Dict[str, Any]:
    """Low-level MCP handshake diagnostics to troubleshoot 400 responses."""
    url = settings.mcp_server_url
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json, text/event-stream",
    }
    if settings.mcp_bearer_token:
        headers["Authorization"] = f"Bearer {settings.mcp_bearer_token}"

    init_payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "initialize",
        "params": {
            "protocolVersion": "2025-06-18",
            "capabilities": {},
            "clientInfo": {"name": "archi-local-chatbot-debug", "version": "0.1.0"},
        },
    }

    out: Dict[str, Any] = {
        "server_url": url,
        "has_bearer_token": bool(settings.mcp_bearer_token),
        "token_length": len(settings.mcp_bearer_token or ""),
    }

    with httpx.Client(timeout=settings.request_timeout_seconds) as client:
        init_resp = client.post(url, headers=headers, json=init_payload)
        sid = init_resp.headers.get("mcp-session-id")
        out["initialize"] = {
            "status_code": init_resp.status_code,
            "session_header": sid,
            "body": init_resp.text[:2500],
        }

        list_headers = dict(headers)
        if sid:
            list_headers["mcp-session-id"] = sid
        tools_payload = {"jsonrpc": "2.0", "id": 2, "method": "tools/list"}
        tools_resp = client.post(url, headers=list_headers, json=tools_payload)
        out["tools_list"] = {
            "status_code": tools_resp.status_code,
            "session_sent": bool(sid),
            "body": tools_resp.text[:2500],
        }

    return out


@app.post("/api/chat", response_model=ChatResponse)
def chat(payload: ChatRequest) -> ChatResponse:
    history = _history_from_payload(payload)
    try:
        answer, used_tools = chat_service.chat(
            history=history,
            message=payload.message,
            system_prompt=payload.system_prompt,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return ChatResponse(answer=answer, used_tools=used_tools)


@app.post("/api/chat/trace", response_model=ChatTraceResponse)
def chat_trace(payload: ChatRequest) -> ChatTraceResponse:
    history = _history_from_payload(payload)
    try:
        answer, used_tools, trace = chat_service.chat_with_trace(
            history=history,
            message=payload.message,
            system_prompt=payload.system_prompt,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return ChatTraceResponse(answer=answer, used_tools=used_tools, trace=trace)


@app.post("/api/chat/stream")
def chat_stream(payload: ChatStreamRequest) -> StreamingResponse:
    history = _history_from_payload(payload)

    def event_stream():
        work_q: Queue[tuple[str, Any]] = Queue()

        def worker() -> None:
            try:
                answer, used_tools, trace = chat_service.chat_with_trace(
                    history=history,
                    message=payload.message,
                    system_prompt=payload.system_prompt,
                )
                work_q.put(("result", {"answer": answer, "used_tools": used_tools, "trace": trace}))
            except Exception as exc:  # noqa: BLE001
                work_q.put(("error", {"detail": str(exc)}))
            finally:
                work_q.put(("end", None))

        Thread(target=worker, daemon=True).start()
        stream_id = str(uuid4())
        yield _sse("start", {"stream_id": stream_id, "started_at": _now_iso()})

        is_done = False
        while not is_done:
            try:
                kind, payload_data = work_q.get(timeout=1.0)
            except Empty:
                yield ": keep-alive\n\n"
                continue

            if kind == "result":
                answer = str(payload_data.get("answer", ""))
                used_tools = payload_data.get("used_tools", [])
                trace = payload_data.get("trace", {})

                if answer:
                    for index, chunk in enumerate(_chunk_text(answer, payload.stream_chunk_chars)):
                        yield _sse("delta", {"index": index, "content": chunk})

                if payload.include_trace:
                    yield _sse("trace", {"used_tools": used_tools, "trace": trace})
                else:
                    yield _sse("tools", {"used_tools": used_tools})

                yield _sse("done", {"answer": answer, "used_tools": used_tools})
            elif kind == "error":
                yield _sse("error", payload_data)
            elif kind == "end":
                is_done = True

        yield _sse("close", {"stream_id": stream_id, "completed_at": _now_iso()})

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/actions/business-process-upload", response_model=ActionResponse)
async def business_process_upload_action(
    file: UploadFile = File(...),
    view_name: str | None = Form(default=None),
) -> ActionResponse:
    text, source_name = await _extract_text_from_upload(file)
    try:
        execution = chat_service.run_business_process_automation(
            content_text=text,
            source_name=source_name,
            view_name=view_name,
        )
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Business process automation failed: {exc}") from exc

    steps_processed = int(execution.get("steps_processed", 0))
    steps_truncated = bool(execution.get("steps_truncated", False))
    status = "partial" if steps_truncated else "ok"
    summary = (
        f"Processed {steps_processed} process steps from '{source_name}'. "
        f"View '{execution.get('view_name', view_name or 'Business Process View')}' was updated in one automation run. "
        "If Archi approval mode is enabled, approve the proposal to apply changes."
    )
    if steps_truncated:
        summary += f" Steps were truncated to the configured maximum of {settings.max_action_steps}."

    return ActionResponse(
        action="business-process-upload",
        status=status,
        summary=summary,
        view_name=str(execution.get("view_name", view_name or "Business Process View")),
        view_id=execution.get("view_id"),
        steps_processed=steps_processed,
        steps_truncated=steps_truncated,
        created_elements=int(execution.get("created_elements", 0)),
        created_relationships=int(execution.get("created_relationships", 0)),
        added_to_view=int(execution.get("added_to_view", 0)),
        added_connections=int(execution.get("added_connections", 0)),
        used_tools=list(execution.get("used_tools", [])),
    )


@app.post("/api/actions/requirements-upload", response_model=ActionResponse)
async def requirements_upload_action(
    file: UploadFile = File(...),
    view_name: str | None = Form(default=None),
) -> ActionResponse:
    text, source_name = await _extract_text_from_upload(file)
    try:
        execution = chat_service.run_requirements_automation(
            content_text=text,
            source_name=source_name,
            view_name=view_name,
        )
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Requirements automation failed: {exc}") from exc

    steps_processed = int(execution.get("steps_processed", 0))
    steps_truncated = bool(execution.get("steps_truncated", False))
    status = "partial" if steps_truncated else "ok"
    summary = (
        f"Generated product architecture from '{source_name}' and updated view "
        f"'{execution.get('view_name', view_name or 'Product Architecture')}' in one automation run. "
        "If Archi approval mode is enabled, approve the proposal to apply changes."
    )
    if steps_truncated:
        summary += " Input was truncated to fit the configured action limits."

    return ActionResponse(
        action="requirements-upload",
        status=status,
        summary=summary,
        view_name=str(execution.get("view_name", view_name or "Product Architecture")),
        view_id=execution.get("view_id"),
        steps_processed=steps_processed,
        steps_truncated=steps_truncated,
        created_elements=int(execution.get("created_elements", 0)),
        created_relationships=int(execution.get("created_relationships", 0)),
        added_to_view=int(execution.get("added_to_view", 0)),
        added_connections=int(execution.get("added_connections", 0)),
        used_tools=list(execution.get("used_tools", [])),
    )


@app.post("/api/actions/business-process-upload/preview", response_model=AutomationPlanResponse)
async def business_process_upload_preview(
    file: UploadFile = File(...),
    view_name: str | None = Form(default=None),
) -> AutomationPlanResponse:
    text, source_name = await _extract_text_from_upload(file)
    try:
        plan = chat_service.plan_business_process_automation(
            content_text=text,
            source_name=source_name,
            view_name=view_name,
        )
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Business process preview failed: {exc}") from exc
    return AutomationPlanResponse(**plan)


@app.post("/api/actions/requirements-upload/preview", response_model=AutomationPlanResponse)
async def requirements_upload_preview(
    file: UploadFile = File(...),
    view_name: str | None = Form(default=None),
) -> AutomationPlanResponse:
    text, source_name = await _extract_text_from_upload(file)
    try:
        plan = chat_service.plan_requirements_automation(
            content_text=text,
            source_name=source_name,
            view_name=view_name,
        )
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Requirements preview failed: {exc}") from exc
    return AutomationPlanResponse(**plan)


@app.post("/api/actions/apply", response_model=ActionResponse)
def apply_plan(payload: ApplyPlanRequest) -> ActionResponse:
    plan_dict = payload.plan.model_dump()
    try:
        execution = chat_service.apply_plan(plan_dict)
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Failed to apply plan: {exc}") from exc

    summary = (
        f"Applied plan: created {execution.get('created_elements', 0)} elements and "
        f"{execution.get('created_relationships', 0)} relationships in view "
        f"'{execution.get('view_name', payload.plan.view_name)}'. "
        "If Archi approval mode is enabled, approve the proposal to apply changes."
    )

    return ActionResponse(
        action=payload.plan.action,
        status="ok",
        summary=summary,
        view_name=str(execution.get("view_name", payload.plan.view_name)),
        view_id=execution.get("view_id"),
        steps_processed=int(execution.get("steps_processed", 0)),
        steps_truncated=False,
        created_elements=int(execution.get("created_elements", 0)),
        created_relationships=int(execution.get("created_relationships", 0)),
        added_to_view=int(execution.get("added_to_view", 0)),
        added_connections=int(execution.get("added_connections", 0)),
        used_tools=list(execution.get("used_tools", [])),
    )


@app.post("/api/conversations/export", response_model=ConversationExportResponse)
def export_conversation(payload: ConversationExportRequest) -> ConversationExportResponse:
    conversation = payload.conversation.model_copy(deep=True)
    conversation, _warnings = _normalize_conversation(conversation)
    now = _now_iso()

    return ConversationExportResponse(
        exported_at=now,
        message_count=len(conversation.history),
        conversation=conversation,
    )


@app.post("/api/conversations/import", response_model=ConversationImportResponse)
def import_conversation(payload: ConversationImportRequest) -> ConversationImportResponse:
    warnings: list[str] = []
    raw_payload = payload.payload

    conversation_payload = raw_payload.get("conversation")
    if isinstance(conversation_payload, dict):
        raw_conversation = conversation_payload
        schema_version = raw_payload.get("schema_version")
        if schema_version and schema_version != "archi-chat-conversation/v1":
            warnings.append(f"Unexpected schema_version '{schema_version}'. Imported with best-effort parsing.")
    else:
        raw_conversation = raw_payload
        warnings.append("Imported payload did not contain an export envelope. Parsed as raw conversation object.")

    try:
        conversation = ConversationArchive.model_validate(raw_conversation)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=f"Invalid conversation payload: {exc}") from exc

    conversation, normalize_warnings = _normalize_conversation(conversation)
    warnings.extend(normalize_warnings)
    return ConversationImportResponse(
        message_count=len(conversation.history),
        warnings=warnings,
        conversation=conversation,
    )
